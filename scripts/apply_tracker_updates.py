from __future__ import annotations

import copy
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "30天重建追蹤.xlsx"
PENDING = ROOT / "pending_updates.json"
RESULT = ROOT / "tracker_update_result.json"

HEADER_ALIASES = {
    "date": ["日期", "date"],
    "meal": ["餐別", "餐次"],
    "food": ["餐點", "內容", "食物"],
    "source": ["來源"],
    "protein": ["蛋白質"],
    "carbs": ["碳水化合物", "碳水", "carbohydrate", "carbs"],
    "calories": ["熱量", "卡路里"],
    "note": ["備註"],
}

DAILY_ALIASES = {
    "date": ["日期", "date"],
    "zero_bet": ["零下注"],
    "viewed_odds": ["查看盤口"],
    "urge": ["衝動 0–10", "衝動0–10", "衝動分數"],
    "trigger": ["觸發原因"],
    "alternative": ["替代行動"],
    "workout_done": ["完成重訓"],
    "walk_minutes": ["散步分鐘"],
    "sleep_hours": ["睡眠小時"],
    "zero_bet_streak": ["連續零下注"],
}

WORKOUT_ALIASES = {
    "date": ["日期", "date"],
    "workout_name": ["訓練名稱"],
    "exercise": ["動作"],
    "set": ["組次", "組數"],
    "weight": ["重量"],
    "reps": ["次數"],
    "dbr": ["dbr", "rpe"],
    "volume": ["訓練量"],
    "note": ["備註"],
}

CARDIO_SHEET = "有氧紀錄"
CARDIO_HEADERS = [
    "日期",
    "有氧類型",
    "段次",
    "時間（分鐘）",
    "坡度（%）",
    "速度（km/h）",
    "距離（km）",
    "備註",
]
CARDIO_ALIASES = {
    "date": ["日期", "date"],
    "activity": ["有氧類型", "類型", "活動"],
    "segment": ["段次", "組次"],
    "minutes": ["時間", "分鐘"],
    "incline": ["坡度"],
    "speed": ["速度"],
    "distance": ["距離"],
    "note": ["備註"],
}


def norm(value):
    return str(value or "").strip().lower()


def scalar_equal(actual, expected):
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return abs(float(actual) - float(expected)) < 1e-9
    return str(actual) == str(expected)


def find_sheet(wb, keyword: str):
    for name in wb.sheetnames:
        if keyword in name:
            return wb[name]
    raise RuntimeError(f"找不到包含「{keyword}」的工作表")


def locate_headers(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        values = {col: norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        mapping = {}
        for field, aliases in HEADER_ALIASES.items():
            for col, value in values.items():
                if any(alias.lower() == value or alias.lower() in value for alias in aliases):
                    mapping[field] = col
                    break
        if len(mapping) >= 3 and "date" in mapping:
            return row, mapping
    raise RuntimeError(f"無法辨識工作表 {ws.title} 的欄位標題")


def locate_daily_headers(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        values = {col: norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        mapping = {}
        for field, aliases in DAILY_ALIASES.items():
            for col, value in values.items():
                if any(alias.lower() == value or alias.lower() in value for alias in aliases):
                    mapping[field] = col
                    break
        if "date" in mapping and len(mapping) >= 6:
            return row, mapping
    raise RuntimeError(f"無法辨識工作表 {ws.title} 的每日檢核欄位")


def locate_workout_headers(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        values = {col: norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        mapping = {}
        for field, aliases in WORKOUT_ALIASES.items():
            for col, value in values.items():
                if any(alias.lower() == value or alias.lower() in value for alias in aliases):
                    mapping[field] = col
                    break
        required = {"date", "exercise", "set", "weight", "reps", "dbr", "volume"}
        if required.issubset(mapping):
            return row, mapping
    raise RuntimeError(f"無法辨識工作表 {ws.title} 的重訓欄位")


def locate_cardio_headers(ws):
    for row in range(1, min(ws.max_row, 8) + 1):
        values = {col: norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        mapping = {}
        for field, aliases in CARDIO_ALIASES.items():
            for col, value in values.items():
                if any(alias.lower() == value or alias.lower() in value for alias in aliases):
                    mapping[field] = col
                    break
        required = {"date", "activity", "segment", "minutes", "incline", "speed"}
        if required.issubset(mapping):
            return row, mapping
    raise RuntimeError(f"無法辨識工作表 {ws.title} 的有氧欄位")


def date_key(value):
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip().replace("/", "-").replace(".", "-")
    try:
        return datetime.fromisoformat(text[:10]).strftime("%Y-%m-%d")
    except ValueError:
        return text


def copy_cell_style(src, dst):
    if src.has_style:
        dst._style = copy.copy(src._style)
    dst.number_format = src.number_format
    dst.alignment = copy.copy(src.alignment)
    dst.fill = copy.copy(src.fill)
    dst.font = copy.copy(src.font)
    dst.border = copy.copy(src.border)
    dst.protection = copy.copy(src.protection)


def copy_row_style(ws, src_row: int, dst_row: int):
    if src_row < 1:
        return
    for col in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def ensure_carbs_column(ws):
    header_row, mapping = locate_headers(ws)
    if "carbs" in mapping:
        return False, mapping["carbs"]
    new_col = ws.max_column + 1
    style_col = mapping.get("protein") or mapping.get("calories") or max(1, new_col - 1)
    ws.cell(header_row, new_col).value = "碳水（g）"
    copy_cell_style(ws.cell(header_row, style_col), ws.cell(header_row, new_col))
    for row in range(header_row + 1, ws.max_row + 1):
        copy_cell_style(ws.cell(row, style_col), ws.cell(row, new_col))
    ws.column_dimensions[get_column_letter(new_col)].width = 12
    return True, new_col


def ensure_cardio_sheet(wb):
    if CARDIO_SHEET in wb.sheetnames:
        return wb[CARDIO_SHEET], False
    ws = wb.create_sheet(CARDIO_SHEET)
    for col, header in enumerate(CARDIO_HEADERS, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    widths = {1: 14, 2: 18, 3: 8, 4: 14, 5: 12, 6: 14, 7: 12, 8: 48}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:H1"
    return ws, True


def append_note(existing, extra):
    existing_text = str(existing or "").strip()
    extra_text = str(extra or "").strip()
    if not extra_text or extra_text in existing_text:
        return existing_text
    return f"{existing_text} {extra_text}".strip()


def food_exists(ws, header_row, mapping, item):
    target_date = item["date"]
    target_meal = norm(item.get("meal"))
    target_food = norm(item.get("food"))
    for row in range(header_row + 1, ws.max_row + 1):
        date = date_key(ws.cell(row, mapping["date"]).value)
        meal = norm(ws.cell(row, mapping["meal"]).value) if mapping.get("meal") else ""
        food = norm(ws.cell(row, mapping["food"]).value) if mapping.get("food") else ""
        if date == target_date and meal == target_meal and food == target_food:
            return row
    return None


def food_matches_exact(ws, row, mapping, item):
    if date_key(ws.cell(row, mapping["date"]).value) != item["date"]:
        return False
    if mapping.get("meal") and norm(ws.cell(row, mapping["meal"]).value) != norm(item.get("meal")):
        return False
    if mapping.get("food") and norm(ws.cell(row, mapping["food"]).value) != norm(item.get("food")):
        return False
    for field in ("protein", "carbs", "calories", "source"):
        if field in item and mapping.get(field) and not scalar_equal(ws.cell(row, mapping[field]).value, item[field]):
            return False
    return True


def find_food_row_exact(ws, header_row, mapping, item):
    for row in range(header_row + 1, ws.max_row + 1):
        if food_matches_exact(ws, row, mapping, item):
            return row
    return None


def apply_food(ws, item):
    header_row, mapping = locate_headers(ws)
    existing = food_exists(ws, header_row, mapping, item)
    if existing and item.get("update_existing"):
        changed_fields = []
        for field in ("source", "protein", "carbs", "calories", "note"):
            if field not in item or not mapping.get(field):
                continue
            if not scalar_equal(ws.cell(existing, mapping[field]).value, item[field]):
                ws.cell(existing, mapping[field]).value = item[field]
                changed_fields.append(field)
        if item.get("note_append") and mapping.get("note"):
            cell = ws.cell(existing, mapping["note"])
            updated = append_note(cell.value, item["note_append"])
            if updated != str(cell.value or ""):
                cell.value = updated
                changed_fields.append("note_append")
        return {"status": "updated" if changed_fields else "already_present", "row": existing, "fields": changed_fields, "item": item}
    if existing:
        return {"status": "already_present", "row": existing, "item": item}
    row = ws.max_row + 1
    copy_row_style(ws, max(header_row + 1, ws.max_row), row)
    values = {
        "date": datetime.strptime(item["date"], "%Y-%m-%d").date(),
        "meal": item.get("meal", ""),
        "food": item.get("food", ""),
        "source": item.get("source", ""),
        "protein": item.get("protein", ""),
        "carbs": item.get("carbs", ""),
        "calories": item.get("calories", ""),
        "note": item.get("note", ""),
    }
    for field, value in values.items():
        if mapping.get(field):
            ws.cell(row, mapping[field]).value = value
    return {"status": "written", "row": row, "item": item}


def apply_delete_food(ws, item):
    header_row, mapping = locate_headers(ws)
    row = find_food_row_exact(ws, header_row, mapping, item)
    if not row:
        return {"status": "already_absent", "row": None, "item": item}
    ws.delete_rows(row, 1)
    return {"status": "deleted", "row": row, "item": item}


def find_daily_row(ws, header_row, mapping, target_date):
    for row in range(header_row + 1, ws.max_row + 1):
        if date_key(ws.cell(row, mapping["date"]).value) == target_date:
            return row
    return None


def apply_daily_checkin(ws, item):
    header_row, mapping = locate_daily_headers(ws)
    row = find_daily_row(ws, header_row, mapping, item["date"])
    if not row:
        raise RuntimeError(f"每日檢核找不到日期 {item['date']}")
    changed_fields = []
    for field in ("zero_bet", "viewed_odds", "urge", "trigger", "alternative", "workout_done", "walk_minutes", "sleep_hours"):
        if field not in item:
            continue
        col = mapping.get(field)
        if not col:
            raise RuntimeError(f"每日檢核找不到欄位 {field}")
        if not scalar_equal(ws.cell(row, col).value, item[field]):
            ws.cell(row, col).value = item[field]
            changed_fields.append(field)
    return {"status": "updated" if changed_fields else "already_present", "row": row, "fields": changed_fields, "item": item}


def workout_exists(ws, header_row, mapping, item):
    target_date = item["date"]
    target_exercise = norm(item.get("exercise"))
    target_set = item.get("set")
    for row in range(header_row + 1, ws.max_row + 1):
        if date_key(ws.cell(row, mapping["date"]).value) != target_date:
            continue
        if norm(ws.cell(row, mapping["exercise"]).value) != target_exercise:
            continue
        if not scalar_equal(ws.cell(row, mapping["set"]).value, target_set):
            continue
        return row
    return None


def find_next_workout_row(ws, header_row, mapping):
    for row in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row, mapping["date"]).value in (None, "") and ws.cell(row, mapping["exercise"]).value in (None, ""):
            return row
    row = ws.max_row + 1
    copy_row_style(ws, max(header_row + 1, ws.max_row), row)
    return row


def set_volume_formula(ws, row, mapping):
    weight_col = get_column_letter(mapping["weight"])
    reps_col = get_column_letter(mapping["reps"])
    ws.cell(row, mapping["volume"]).value = f'=IF(OR({weight_col}{row}="",{reps_col}{row}=""),"",{weight_col}{row}*{reps_col}{row})'


def item_dbr(item):
    if "dbr" in item:
        return item.get("dbr")
    return item.get("rpe", "")


def apply_workout_set(ws, item):
    header_row, mapping = locate_workout_headers(ws)
    existing = workout_exists(ws, header_row, mapping, item)
    if existing and item.get("update_existing"):
        changed_fields = []
        field_values = {
            "workout_name": item.get("workout_name", "重訓"),
            "weight": item.get("weight") if "weight" in item else None,
            "reps": item.get("reps") if "reps" in item else None,
            "dbr": item_dbr(item) if ("dbr" in item or "rpe" in item) else None,
            "note": item.get("note") if "note" in item else None,
        }
        for field, expected in field_values.items():
            if expected is None or not mapping.get(field):
                continue
            if not scalar_equal(ws.cell(existing, mapping[field]).value, expected):
                ws.cell(existing, mapping[field]).value = expected
                changed_fields.append(field)
        if item.get("note_append") and mapping.get("note"):
            cell = ws.cell(existing, mapping["note"])
            updated = append_note(cell.value, item["note_append"])
            if updated != str(cell.value or ""):
                cell.value = updated
                changed_fields.append("note_append")
        set_volume_formula(ws, existing, mapping)
        return {"status": "updated" if changed_fields else "already_present", "row": existing, "fields": changed_fields, "item": item}
    if existing:
        return {"status": "already_present", "row": existing, "item": item}
    row = find_next_workout_row(ws, header_row, mapping)
    values = {
        "date": datetime.strptime(item["date"], "%Y-%m-%d").date(),
        "workout_name": item.get("workout_name", "重訓"),
        "exercise": item.get("exercise", ""),
        "set": item.get("set", ""),
        "weight": item.get("weight", ""),
        "reps": item.get("reps", ""),
        "dbr": item_dbr(item),
        "note": item.get("note", ""),
    }
    for field, value in values.items():
        if mapping.get(field):
            ws.cell(row, mapping[field]).value = value
    set_volume_formula(ws, row, mapping)
    return {"status": "written", "row": row, "item": item}


def cardio_exists(ws, header_row, mapping, item):
    for row in range(header_row + 1, ws.max_row + 1):
        if date_key(ws.cell(row, mapping["date"]).value) != item["date"]:
            continue
        if norm(ws.cell(row, mapping["activity"]).value) != norm(item.get("activity")):
            continue
        if not scalar_equal(ws.cell(row, mapping["segment"]).value, item.get("segment", 1)):
            continue
        return row
    return None


def apply_cardio(ws, item):
    header_row, mapping = locate_cardio_headers(ws)
    existing = cardio_exists(ws, header_row, mapping, item)
    values = {
        "date": datetime.strptime(item["date"], "%Y-%m-%d").date(),
        "activity": item.get("activity", "跑步機爬坡"),
        "segment": item.get("segment", 1),
        "minutes": item.get("minutes", ""),
        "incline": item.get("incline", ""),
        "speed": item.get("speed", ""),
        "distance": item.get("distance", ""),
        "note": item.get("note", ""),
    }
    if existing and item.get("update_existing"):
        changed_fields = []
        for field, expected in values.items():
            if field == "date" or field not in item and field not in {"activity", "segment"}:
                continue
            if mapping.get(field) and not scalar_equal(ws.cell(existing, mapping[field]).value, expected):
                ws.cell(existing, mapping[field]).value = expected
                changed_fields.append(field)
        if item.get("note_append") and mapping.get("note"):
            cell = ws.cell(existing, mapping["note"])
            updated = append_note(cell.value, item["note_append"])
            if updated != str(cell.value or ""):
                cell.value = updated
                changed_fields.append("note_append")
        return {"status": "updated" if changed_fields else "already_present", "row": existing, "fields": changed_fields, "item": item}
    if existing:
        return {"status": "already_present", "row": existing, "item": item}
    row = ws.max_row + 1
    for field, value in values.items():
        if mapping.get(field):
            ws.cell(row, mapping[field]).value = value
    return {"status": "written", "row": row, "item": item}


def verify_food(ws, result):
    item = result["item"]
    header_row, mapping = locate_headers(ws)
    row = food_exists(ws, header_row, mapping, item)
    if not row:
        raise RuntimeError(f"驗證失敗：找不到餐點 {item}")
    for field in ("protein", "carbs", "calories"):
        if field not in item or not mapping.get(field) or item[field] in (None, ""):
            continue
        if not scalar_equal(ws.cell(row, mapping[field]).value, item[field]):
            raise RuntimeError(f"驗證失敗：{field} 預期 {item[field]}，實際 {ws.cell(row, mapping[field]).value}")
    if item.get("note_append") and mapping.get("note") and item["note_append"] not in str(ws.cell(row, mapping["note"]).value or ""):
        raise RuntimeError("驗證失敗：備註修正未寫入")
    return {"verified": True, "row": row, "date": item["date"], "meal": item.get("meal"), "food": item.get("food"), "status": result.get("status")}


def verify_delete_food(ws, result):
    item = result["item"]
    header_row, mapping = locate_headers(ws)
    row = find_food_row_exact(ws, header_row, mapping, item)
    if row:
        raise RuntimeError(f"驗證失敗：欲刪除餐點仍存在於第 {row} 列：{item}")
    return {"verified": True, "date": item["date"], "meal": item.get("meal"), "food": item.get("food"), "status": result.get("status")}


def verify_daily_checkin(ws, result):
    item = result["item"]
    header_row, mapping = locate_daily_headers(ws)
    row = find_daily_row(ws, header_row, mapping, item["date"])
    if not row:
        raise RuntimeError(f"驗證失敗：找不到每日檢核日期 {item['date']}")
    for field in ("zero_bet", "viewed_odds", "urge", "trigger", "alternative", "workout_done", "walk_minutes", "sleep_hours"):
        if field in item and not scalar_equal(ws.cell(row, mapping[field]).value, item[field]):
            raise RuntimeError(f"驗證失敗：{field} 預期 {item[field]}，實際 {ws.cell(row, mapping[field]).value}")
    streak_col = mapping.get("zero_bet_streak")
    if streak_col:
        formula = ws.cell(row, streak_col).value
        if not isinstance(formula, str) or not formula.startswith("="):
            raise RuntimeError("驗證失敗：連續零下注公式遺失")
    return {"verified": True, "row": row, "date": item["date"], "status": result.get("status")}


def verify_workout_set(ws, result):
    item = result["item"]
    header_row, mapping = locate_workout_headers(ws)
    row = workout_exists(ws, header_row, mapping, item)
    if not row:
        raise RuntimeError(f"驗證失敗：找不到重訓組 {item}")
    checks = {"weight": item.get("weight") if "weight" in item else None, "reps": item.get("reps") if "reps" in item else None}
    if "dbr" in item or "rpe" in item:
        checks["dbr"] = item_dbr(item)
    for field, expected in checks.items():
        if expected is None:
            continue
        actual = ws.cell(row, mapping[field]).value
        if expected in (None, ""):
            if actual not in (None, ""):
                raise RuntimeError(f"驗證失敗：{field} 應為空白，實際 {actual}")
        elif not scalar_equal(actual, expected):
            raise RuntimeError(f"驗證失敗：{field} 預期 {expected}，實際 {actual}")
    if item.get("note_append") and mapping.get("note") and item["note_append"] not in str(ws.cell(row, mapping["note"]).value or ""):
        raise RuntimeError("驗證失敗：重訓備註追加未寫入")
    if "note" in item and not item.get("note_append") and mapping.get("note") and str(ws.cell(row, mapping["note"]).value or "") != str(item.get("note") or ""):
        raise RuntimeError("驗證失敗：重訓備註不一致")
    formula = ws.cell(row, mapping["volume"]).value
    if not isinstance(formula, str) or not formula.startswith("="):
        raise RuntimeError("驗證失敗：訓練量公式遺失")
    return {"verified": True, "row": row, "date": item["date"], "exercise": item.get("exercise"), "set": item.get("set"), "status": result.get("status")}


def verify_cardio(ws, result):
    item = result["item"]
    header_row, mapping = locate_cardio_headers(ws)
    row = cardio_exists(ws, header_row, mapping, item)
    if not row:
        raise RuntimeError(f"驗證失敗：找不到有氧紀錄 {item}")
    for field in ("minutes", "incline", "speed", "distance"):
        if field not in item:
            continue
        if not scalar_equal(ws.cell(row, mapping[field]).value, item[field]):
            raise RuntimeError(f"驗證失敗：有氧 {field} 預期 {item[field]}，實際 {ws.cell(row, mapping[field]).value}")
    if item.get("note_append") and mapping.get("note") and item["note_append"] not in str(ws.cell(row, mapping["note"]).value or ""):
        raise RuntimeError("驗證失敗：有氧備註追加未寫入")
    return {"verified": True, "row": row, "date": item["date"], "activity": item.get("activity"), "segment": item.get("segment", 1), "status": result.get("status")}


def main():
    payload = json.loads(PENDING.read_text(encoding="utf-8"))
    updates = payload.get("updates", [])

    wb = load_workbook(WORKBOOK)
    food_ws = find_sheet(wb, "飲食")
    daily_ws = find_sheet(wb, "每日檢核")
    workout_ws = find_sheet(wb, "重訓")
    cardio_ws, cardio_created = ensure_cardio_sheet(wb)
    added_column, carbs_col = ensure_carbs_column(food_ws)

    results = []
    for item in updates:
        kind = item.get("type")
        if kind == "food":
            results.append(apply_food(food_ws, item))
        elif kind == "delete_food":
            results.append(apply_delete_food(food_ws, item))
        elif kind == "daily_checkin":
            results.append(apply_daily_checkin(daily_ws, item))
        elif kind == "workout_set":
            results.append(apply_workout_set(workout_ws, item))
        elif kind == "cardio":
            results.append(apply_cardio(cardio_ws, item))
        else:
            raise RuntimeError(f"尚未支援的 update type: {kind}")

    changed = cardio_created or added_column or any(r.get("status") in {"written", "updated", "deleted"} for r in results)
    if changed:
        wb.save(WORKBOOK)

    verify_wb = load_workbook(WORKBOOK, data_only=False)
    verify_food_ws = find_sheet(verify_wb, "飲食")
    verify_daily_ws = find_sheet(verify_wb, "每日檢核")
    verify_workout_ws = find_sheet(verify_wb, "重訓")
    verify_cardio_ws = find_sheet(verify_wb, "有氧")
    _, verify_mapping = locate_headers(verify_food_ws)
    if "carbs" not in verify_mapping:
        raise RuntimeError("驗證失敗：飲食紀錄沒有碳水欄位")

    verifications = []
    for result in results:
        kind = result["item"].get("type")
        if kind == "food":
            verifications.append(verify_food(verify_food_ws, result))
        elif kind == "delete_food":
            verifications.append(verify_delete_food(verify_food_ws, result))
        elif kind == "daily_checkin":
            verifications.append(verify_daily_checkin(verify_daily_ws, result))
        elif kind == "workout_set":
            verifications.append(verify_workout_set(verify_workout_ws, result))
        elif kind == "cardio":
            verifications.append(verify_cardio(verify_cardio_ws, result))

    output = {
        "migration": {
            "carbs_column_added": added_column,
            "carbs_column": carbs_col,
            "cardio_sheet_created": cardio_created,
        },
        "results": results,
        "verification": verifications,
    }
    RESULT.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if updates:
        PENDING.write_text(json.dumps({"updates": []}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(output, ensure_ascii=False))


if __name__ == "__main__":
    main()
