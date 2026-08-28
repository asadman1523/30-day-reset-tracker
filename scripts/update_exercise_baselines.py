from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "30天重建追蹤.xlsx"
SHEET_NAME = "動作基準"
START_DATE = "2026-08-01"
HEADERS = [
    "動作",
    "建議下次重量（kg）",
    "建議目標次數",
    "建議DBR",
    "最近訓練日期",
    "最近使用重量（kg）",
    "最近次數",
    "最近DBR",
    "最近訓練摘要",
    "狀態／注意",
    "建議依據",
]

PAIN_RE = re.compile(r"痛|疼|不適|刺痛|骨盆|薦髂", re.IGNORECASE)


def norm(value):
    return str(value or "").strip()


def date_key(value):
    if value in (None, ""):
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip().replace("/", "-").replace(".", "-")
    try:
        return datetime.fromisoformat(text[:10]).strftime("%Y-%m-%d")
    except ValueError:
        return text


def as_number(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip()
        return float(text) if text else None
    except (TypeError, ValueError):
        return None


def tidy_number(value):
    if value is None:
        return ""
    value = float(value)
    return int(value) if value.is_integer() else value


def find_workout_sheet(wb):
    for name in wb.sheetnames:
        if "重訓" in name:
            return wb[name]
    raise RuntimeError("找不到重訓工作表")


def detect_headers(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        values = {col: norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        mapping = {}
        for col, value in values.items():
            upper = value.upper()
            if "日期" in value:
                mapping["date"] = col
            elif "動作" in value:
                mapping["exercise"] = col
            elif "重量" in value:
                mapping["weight"] = col
            elif "次數" in value:
                mapping["reps"] = col
            elif upper in {"DBR", "RPE"}:
                mapping["dbr"] = col
            elif "組次" in value or "組數" in value:
                mapping["set"] = col
            elif "備註" in value:
                mapping["note"] = col
        if {"date", "exercise", "weight", "reps", "dbr"}.issubset(mapping):
            return row, mapping
    raise RuntimeError("無法辨識重訓紀錄欄位")


def group_weight_rows(rows):
    grouped = defaultdict(list)
    for row in rows:
        weight = row["weight"]
        if weight is not None and weight > 0:
            grouped[weight].append(row)
    return grouped


def choose_suggested_weight(rows):
    grouped = group_weight_rows(rows)
    if not grouped:
        return 0.0, "自重動作；下次仍以自重為起點。"

    # 不把「最重」直接等同「適合下次」。先找能穩定落在有效次數與 DBR 的工作重量。
    scored = []
    for weight, sets in grouped.items():
        productive = [
            s for s in sets
            if s["dbr"] is not None and 1 <= s["dbr"] <= 3
            and s["reps"] is not None and 8 <= s["reps"] <= 20
        ]
        nonfailure = [s for s in sets if s["dbr"] is not None and s["dbr"] >= 1]
        failures = [s for s in sets if s["dbr"] is not None and s["dbr"] == 0]
        known = [s for s in sets if s["dbr"] is not None]
        avg_dbr = sum(s["dbr"] for s in known) / len(known) if known else None
        score = (
            len(productive),
            len(nonfailure),
            -len(failures),
            len(sets),
            -(abs(avg_dbr - 2) if avg_dbr is not None else 99),
            weight,
        )
        scored.append((score, weight, sets, productive, nonfailure, failures))

    best = max(scored, key=lambda x: x[0])
    _, weight, sets, productive, nonfailure, failures = best
    if productive:
        reason = (
            f"最近一次訓練中 {weight:g} kg 有 {len(productive)} 組落在約 8–20 下、DBR1–3；"
            "以可重複的工作組表現優先，不直接選最高重量。"
        )
        return weight, reason
    if nonfailure:
        reason = (
            f"最近一次訓練缺少完整的理想工作組，但 {weight:g} kg 的非力竭組最穩定；"
            "下次先延續此重量再依 DBR 調整。"
        )
        return weight, reason

    known_any = any(s["dbr"] is not None for sets in grouped.values() for s in sets)
    if known_any:
        # 若最新一次全部已知組都到 DBR0，採該次最低正重量，避免把力竭重量直接當起始重量。
        weight = min(grouped)
        return weight, "最近一次已知組多為力竭；下次採該次較低重量重新建立 DBR1–3 的工作組。"

    # 完全沒有 DBR 時，採最近一次最常使用的重量；同次數時取較低者，不以最高重量為準。
    counts = Counter({weight: len(sets) for weight, sets in grouped.items()})
    max_count = max(counts.values())
    weight = min(w for w, count in counts.items() if count == max_count)
    return weight, "最近一次沒有 DBR；先採最常使用的重量（同頻率時取較低者），不直接採最高重量。"


def summarize_latest(rows):
    parts = []
    for row in rows:
        weight = "自重" if row["weight"] in (None, 0) else f"{row['weight']:g}kg"
        reps = "?" if row["reps"] is None else f"{row['reps']:g}"
        dbr = "?" if row["dbr"] is None else f"{row['dbr']:g}"
        side = ""
        note = row.get("note", "")
        if "左" in note:
            side = "左"
        elif "右" in note:
            side = "右"
        parts.append(f"{side}{weight}×{reps}/DBR{dbr}")
    return "；".join(parts)


def main():
    wb = load_workbook(WORKBOOK)
    ws = find_workout_sheet(wb)
    header_row, mapping = detect_headers(ws)

    by_exercise = defaultdict(list)
    for row in range(header_row + 1, ws.max_row + 1):
        exercise = norm(ws.cell(row, mapping["exercise"]).value)
        dt = date_key(ws.cell(row, mapping["date"]).value)
        if not exercise or not dt or dt < START_DATE:
            continue
        by_exercise[exercise].append({
            "row": row,
            "date": dt,
            "set": as_number(ws.cell(row, mapping.get("set", 0)).value) if mapping.get("set") else None,
            "weight": as_number(ws.cell(row, mapping["weight"]).value),
            "reps": as_number(ws.cell(row, mapping["reps"]).value),
            "dbr": as_number(ws.cell(row, mapping["dbr"]).value),
            "note": norm(ws.cell(row, mapping["note"]).value) if mapping.get("note") else "",
        })

    if SHEET_NAME in wb.sheetnames:
        base = wb[SHEET_NAME]
        wb.remove(base)
    base = wb.create_sheet(SHEET_NAME)

    for col, header in enumerate(HEADERS, 1):
        cell = base.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    out_row = 2
    for exercise in sorted(by_exercise):
        all_rows = by_exercise[exercise]
        latest_date = max(r["date"] for r in all_rows)
        latest = [r for r in all_rows if r["date"] == latest_date]
        latest.sort(key=lambda r: ((r["set"] if r["set"] is not None else 9999), r["row"]))

        suggested, reason = choose_suggested_weight(latest)
        latest_notes = " ".join(r.get("note", "") for r in latest)
        pain_flag = bool(PAIN_RE.search(latest_notes))

        recent = latest[-1]
        status = "正常"
        if pain_flag:
            status = "疼痛／不適：暫不自動推薦重量"
            suggested_display = ""
            reason = "最近一次訓練備註出現疼痛／不適；保留歷史重量，但下次先確認無痛，不依重量基準直接加載。"
        else:
            suggested_display = tidy_number(suggested)

        base.cell(out_row, 1, exercise)
        base.cell(out_row, 2, suggested_display)
        base.cell(out_row, 3, "10–15（特殊動作可依既有表現調整）")
        base.cell(out_row, 4, "1–3；最後一組可視情況 DBR0")
        base.cell(out_row, 5, latest_date)
        base.cell(out_row, 6, tidy_number(recent["weight"]))
        base.cell(out_row, 7, tidy_number(recent["reps"]))
        base.cell(out_row, 8, tidy_number(recent["dbr"]))
        base.cell(out_row, 9, summarize_latest(latest))
        base.cell(out_row, 10, status)
        base.cell(out_row, 11, reason)
        out_row += 1

    widths = {
        1: 22, 2: 20, 3: 30, 4: 28, 5: 15, 6: 20, 7: 12, 8: 12,
        9: 80, 10: 30, 11: 75,
    }
    for col, width in widths.items():
        base.column_dimensions[get_column_letter(col)].width = width
    for row in base.iter_rows(min_row=2, max_row=max(2, out_row - 1), min_col=9, max_col=11):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    base.freeze_panes = "A2"
    base.auto_filter.ref = f"A1:K{max(1, out_row - 1)}"

    wb.save(WORKBOOK)
    print(f"Updated {SHEET_NAME}: {len(by_exercise)} exercises scanned from {START_DATE}")


if __name__ == "__main__":
    main()
