from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "30天重建追蹤.xlsx"
MARKER = "DBR_SCHEMA_MIGRATED"


def find_workout_sheet(wb):
    for name in wb.sheetnames:
        if "重訓" in name:
            return wb[name]
    raise RuntimeError("找不到重訓工作表")


def find_effort_header(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, ws.max_column + 1):
            value = str(ws.cell(row, col).value or "").strip().upper()
            if value in {"RPE", "DBR"}:
                return row, col, value
    raise RuntimeError("找不到 RPE/DBR 欄位")


def prepare():
    wb = load_workbook(WORKBOOK)
    ws = find_workout_sheet(wb)
    row, col, header = find_effort_header(ws)
    cell = ws.cell(row, col)

    # The existing updater still locates this column by the legacy RPE header.
    # Temporarily expose it as RPE, but tag it so finalize knows the values
    # are already DBR and must not be converted again.
    if header == "DBR":
        cell.value = "RPE"
        cell.comment = Comment(MARKER, "tracker")
        wb.save(WORKBOOK)


def parse_dbr_from_note(note):
    match = re.search(r"\bDBR\s*([0-9]+(?:\.[0-9]+)?)\b", str(note or ""), re.IGNORECASE)
    if not match:
        return None
    value = float(match.group(1))
    return int(value) if value.is_integer() else value


def finalize():
    wb = load_workbook(WORKBOOK)
    ws = find_workout_sheet(wb)
    header_row, col, header = find_effort_header(ws)
    cell = ws.cell(header_row, col)

    if header == "DBR":
        return

    already_migrated = bool(cell.comment and MARKER in str(cell.comment.text or ""))

    if not already_migrated:
        # Initial migration. Newer rows already stored DBR in 備註, so prefer
        # that exact user-reported value. Older rows with only RPE are converted
        # by DBR = 10 - RPE.
        note_col = None
        for c in range(1, ws.max_column + 1):
            if "備註" in str(ws.cell(header_row, c).value or ""):
                note_col = c
                break

        for r in range(header_row + 1, ws.max_row + 1):
            effort_cell = ws.cell(r, col)
            note = ws.cell(r, note_col).value if note_col else ""
            note_dbr = parse_dbr_from_note(note)
            if note_dbr is not None:
                effort_cell.value = note_dbr
                continue

            value = effort_cell.value
            if isinstance(value, (int, float)):
                converted = 10 - float(value)
                effort_cell.value = int(converted) if converted.is_integer() else converted

    cell.value = "DBR"
    cell.comment = None
    wb.save(WORKBOOK)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("mode", choices=["prepare", "finalize"])
    args = parser.parse_args()
    if args.mode == "prepare":
        prepare()
    else:
        finalize()


if __name__ == "__main__":
    main()
