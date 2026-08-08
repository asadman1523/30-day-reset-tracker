from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "30天重建追蹤.xlsx"
OUTPUT = ROOT / "dashboard-data.json"
TARGET_SHEETS = ("每日檢核", "飲食紀錄", "重訓紀錄")


def serialize(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def detect_header_row(ws):
    best = None
    for row in range(1, min(ws.max_row, 15) + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        labels = [str(v).strip() for v in values if v not in (None, "")]
        score = len(labels)
        has_date = any("日期" in label or label.lower() == "date" for label in labels)
        if has_date and score >= 2:
            return row
        if score and (best is None or score > best[0]):
            best = (score, row)
    return best[1] if best else 1


def unique_headers(ws, header_row):
    used = {}
    headers = []
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, col).value
        base = str(raw).strip() if raw not in (None, "") else f"欄位{col}"
        count = used.get(base, 0) + 1
        used[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def export_sheet(ws):
    header_row = detect_header_row(ws)
    headers = unique_headers(ws, header_row)
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = [serialize(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in values):
            continue
        rows.append({headers[i]: values[i] if values[i] is not None else "" for i in range(len(headers))})
    return {"header_row": header_row, "headers": headers, "rows": rows}


def main():
    wb = load_workbook(WORKBOOK, data_only=False)
    sheets = {name: export_sheet(wb[name]) for name in TARGET_SHEETS if name in wb.sheetnames}
    payload = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "workbook": WORKBOOK.name,
        "sheets": sheets,
    }
    OUTPUT.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT} with {', '.join(sheets)}")


if __name__ == "__main__":
    main()
